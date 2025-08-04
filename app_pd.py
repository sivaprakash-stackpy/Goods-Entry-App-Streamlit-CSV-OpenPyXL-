import streamlit as st
import pandas as pd
import os
from openpyxl import load_workbook

# File path for the data
DATA_FILE = "data/goods_data.csv"

# Create the data directory and file if not exist
os.makedirs("data", exist_ok=True)
if not os.path.exists(DATA_FILE):
    df_init = pd.DataFrame(columns=["Item Name", "Quantity", "Price"])
    df_init.to_csv(DATA_FILE, index=False)

# UI
st.title("🛒 Goods Entry Form")

st.markdown("Fill the details below to add a new item to the inventory")

item_name = st.text_input("Item Name")
quantity = st.number_input("Quantity", min_value=1, step=1)
price = st.number_input("Price (in $)", min_value=0.0, step=0.01, format="%.2f")

if st.button("📥 Submit"):
    if item_name.strip() == "":
        st.warning("Please enter an item name.")
    else:
        new_data = pd.DataFrame([[item_name, int(quantity), float(price)]],
                                columns=["Item Name", "Quantity", "Price"])
        new_data.to_csv(DATA_FILE, mode='a', header=False, index=False)

        # Use openpyxl to format the Excel version (optional)
        excel_file = DATA_FILE.replace(".csv", ".xlsx")
        if not os.path.exists(excel_file):
            new_data.to_excel(excel_file, index=False)
        else:
            # Append using openpyxl
            book = load_workbook(excel_file)
            writer = pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay')
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            start_row = book.active.max_row
            new_data.to_excel(writer, index=False, header=False, startrow=start_row)
            writer.close()

        st.success(f"✅ '{item_name}' added successfully!")

# Display current data
if st.checkbox("📄 Show Current Goods List"):
    df = pd.read_csv(DATA_FILE)
    st.dataframe(df)
