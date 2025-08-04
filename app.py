import streamlit as st
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# File path
DATA_FILE = "data/goods_data.xlsx"

# Ensure data folder exists
os.makedirs("data", exist_ok=True)

# Initialize Excel file if it doesn't exist
if not os.path.exists(DATA_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Goods"
    ws.append(["Item Name", "Quantity", "Price"])
    wb.save(DATA_FILE)

# Streamlit UI
st.title("📦 Goods Entry Form")

item_name = st.text_input("Item Name")
quantity = st.number_input("Quantity", min_value=1, step=1)
price = st.number_input("Price ($)", min_value=0.0, step=0.01, format="%.2f")

if st.button("➕ Add Item"):
    if item_name.strip() == "":
        st.warning("Item name cannot be empty!")
    else:
        # Load workbook and append new data
        wb = load_workbook(DATA_FILE)
        ws = wb["Goods"]
        ws.append([item_name, quantity, price])
        wb.save(DATA_FILE)
        st.success(f"✅ '{item_name}' added successfully.")

# Display existing data
if st.checkbox("📄 Show All Entries"):
    wb = load_workbook(DATA_FILE)
    ws = wb["Goods"]

    # Extract data for display
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(row)

    if data:
        st.write("### Inventory List:")
        st.table(data)
    else:
        st.info("No data available yet.")
